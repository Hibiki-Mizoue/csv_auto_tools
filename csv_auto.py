import os
import shutil
import pandas as pd
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

# =========================
# 入力CSVの決定とコピー
# =========================
candidates = ['examples/sample_data_1.csv', 'examples/sample_data_2.csv']
src = next((p for p in candidates if os.path.exists(p)), None)
if not src:
    raise FileNotFoundError("examples フォルダにサンプルCSVが見つかりません。")

work_csv = 'working_data.csv'
shutil.copy(src, work_csv)
print(f"{src} → {work_csv} にコピーしました")

# =========================
# 読み込み & 集計
# =========================
df = pd.read_csv(work_csv)

# 日別×商品 の数量ピボット
daily_product_count = df.pivot_table(
    index='日付', columns='商品', values='数量', aggfunc='sum', fill_value=0
).reset_index()
daily_product_count.columns.name = None  # 余分な階層名を除去

# 日別売上合計
daily_total = df.groupby('日付', as_index=False)['売上'].sum()

# 商品別売上合計
category_total = df.groupby('商品', as_index=False)['売上'].sum()

# =========================
# 出力（1ブック・複数シート）
# =========================
out_dir = "output"
out_path = os.path.join(out_dir, "report.xlsx")
os.makedirs(out_dir, exist_ok=True)

# pandasのExcelWriterで書き出し（B列・3行目開始 = startcol=1, startrow=2）
writer = pd.ExcelWriter(out_path, engine="openpyxl")
daily_product_count.to_excel(writer, sheet_name="日別商品数量", index=False, startrow=2, startcol=1)
daily_total.to_excel(writer,        sheet_name="日別売上",   index=False, startrow=2, startcol=1)
category_total.to_excel(writer,     sheet_name="商品別売上", index=False, startrow=2, startcol=1)

# =========================
# 書式適用 共通関数
# =========================
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

def style_sheet(ws, title: str, header_fill_color="FFCCFFCC"):
    # 右端の完全空列を削除
    col = ws.max_column
    while col > 1:
        empty = True
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            if v not in (None, ""):
                empty = False
                break
        if empty:
            ws.delete_cols(col)
        col -= 1

    max_col = ws.max_column

    # 列幅の自動調整（最低幅を確保しておく）
    MIN_WIDTH = 12  # 最低列幅
    for c in range(2, max_col + 1):
        letter = get_column_letter(c)
        max_len = 0
        for r in range(3, ws.max_row + 1):
            v = ws.cell(row=r, column=c).value
            L = len(str(v)) if v is not None else 0
            if L > max_len:
                max_len = L
        ws.column_dimensions[letter].width = max(MIN_WIDTH, min(max_len + 2, 40))

    # タイトル（B2～最右列）を結合・配置
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=max_col)
    ws["B2"] = title
    ws["B2"].font = Font(size=14, bold=True)
    # ▼ ここが重要：自動縮小してセル幅に収める／折返ししない
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=False, shrink_to_fit=True)
    # タイトル行の高さも少し確保
    ws.row_dimensions[2].height = 22

    thin = Side(style="thin", color="000000")
    double = Side(style="double", color="000000")

    # ヘッダー装飾（太字・背景色・下二重線・中央寄せ）
    for c in range(2, max_col + 1):
        cell = ws.cell(row=3, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=double)

    # データ枠線 & 売上列の数値書式
    for r in range(4, ws.max_row + 1):
        for c in range(2, max_col + 1):
            cell = ws.cell(row=r, column=c)
            header = ws.cell(row=3, column=c).value
            if isinstance(cell.value, (int, float)) and header and "売上" in str(header):
                cell.number_format = '#,##0'
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# =========================
# 書式適用（各シート）
# =========================
wb = writer.book
ws1 = writer.sheets["日別商品数量"]
ws2 = writer.sheets["日別売上"]
ws3 = writer.sheets["商品別売上"]

style_sheet(ws1, "日別商品数量集計")
style_sheet(ws2, "日別売上集計")
style_sheet(ws3, "カテゴリ別売上集計")

# 保存（開いていると PermissionError になります）
try:
    writer.close()
    print(f"{out_path} に出力しました")
except PermissionError:
    print(f"[エラー] {out_path} が開いているため保存できません。Excel を閉じてから再実行してください。")
